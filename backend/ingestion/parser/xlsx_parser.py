"""XLSX 解析器。

每 sheet → markdown 表格段，sheet 名进 title_tree。
大 sheet 自动按行分段（每段 ≤ 800 字 + 重复表头），避免 chunker 硬切表格行。
"""
from pathlib import Path
from openpyxl import load_workbook
from backend.ingestion.parser.types import ParseResult, TitleNode

# 每段最多字符数（留 200 字 buffer 给 chunker MAX_CHARS=1000，避免硬切）
ROW_GROUP_MAX_CHARS = 800


def _row_to_line(row) -> str:
    return " | ".join(str(c) if c is not None else "" for c in row)


async def parse(path: Path) -> ParseResult:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names: list[str] = []
    parts: list[str] = []

    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        parts.append(f"## Sheet: {ws.title}")
        parts.append("")  # 让 chunker 把标题段和数据段分开

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parts.append("")  # sheet 之间空行
            continue

        # 表头是第一行（即使没真正的表头，也按"第一行"处理，重复出现也无害）
        header_line = _row_to_line(rows[0])

        # 第一段开头先放表头
        current_group: list[str] = [header_line]
        current_chars = len(header_line)

        for row in rows[1:]:
            line = _row_to_line(row)
            # 加这一行会超阈值 → 关上当前段，开新段（新段重复表头）
            if current_chars + len(line) + 1 > ROW_GROUP_MAX_CHARS:
                parts.extend(current_group)
                parts.append("")  # 空行 = chunker 段分隔
                current_group = [header_line]
                current_chars = len(header_line)
            current_group.append(line)
            current_chars += len(line) + 1  # +1 for join 的 \n

        # flush 最后一段
        parts.extend(current_group)
        parts.append("")  # sheet 之间空行

    raw_text = "\n".join(parts)

    # 给每个 sheet 加 TitleNode（level=2，对应 "## Sheet:"）
    title_nodes: list[TitleNode] = []
    cursor = 0
    for sheet_name in sheet_names:
        header = f"## Sheet: {sheet_name}"
        offset = raw_text.find(header, cursor)
        if offset >= 0:
            title_nodes.append(TitleNode(level=2, text=sheet_name, char_offset=offset))
            cursor = offset + len(header)

    return ParseResult(
        raw_text=raw_text,
        title_tree=title_nodes,
        content_type="document",
        metadata={"sheet_names": sheet_names},
    )
